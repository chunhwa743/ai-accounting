"""Structured output an accountant can review and use.

Three artefacts, because "give me the data" means different things at different
points in the work:

  * a review pack, to check the coding line by line
  * journal entries, to post into the accounting system
  * a client query list, to send out and chase

The journal generator is the only place in the system that converts to debits
and credits, and it validates that every entry balances before writing it.
"""

from __future__ import annotations

import csv
import json
from dataclasses import dataclass, field
from decimal import Decimal
from io import StringIO
from pathlib import Path

from .db import Repositories
from .models import AllocationStatus, DecisionMethod
from .reference import get_chart_of_accounts, get_tax_codes

# The contra side of every bank line.
BANK_ACCOUNT_CODE = "090"


@dataclass
class JournalLine:
    date: str
    account_code: str
    account_name: str
    description: str
    debit: Decimal
    credit: Decimal
    tax_code: str | None
    reference: str | None


@dataclass
class JournalEntry:
    entry_id: str
    date: str
    narration: str
    lines: list[JournalLine] = field(default_factory=list)

    @property
    def total_debit(self) -> Decimal:
        return sum((line.debit for line in self.lines), Decimal("0"))

    @property
    def total_credit(self) -> Decimal:
        return sum((line.credit for line in self.lines), Decimal("0"))

    @property
    def balances(self) -> bool:
        return abs(self.total_debit - self.total_credit) <= Decimal("0.01")


def split_tax(gross: Decimal, tax_code: str | None) -> tuple[Decimal, Decimal]:
    """Separate a GST-inclusive amount into net and tax.

    Blocked input tax is deliberately left inside the expense: GST on medical
    costs, private car running costs and club subscriptions cannot be
    reclaimed, so booking it as a receivable from IRAS would overstate assets
    and understate the expense.

    Claimability only governs *purchases*. On a sale the GST still has to come
    out, because it was collected on the government's behalf and is owed to
    them - it was never the client's money.
    """
    codes = get_tax_codes()
    entry = codes.get(tax_code)
    if entry is None or entry.rate == 0:
        return gross, Decimal("0.00")
    if entry.applies_to == "purchase" and not entry.claimable:
        return gross, Decimal("0.00")

    net = (gross / (1 + entry.rate)).quantize(Decimal("0.01"))
    return net, (gross - net).quantize(Decimal("0.01"))


def build_journal(repos: Repositories, run_id: int) -> list[JournalEntry]:
    """One entry per bank transaction, with its allocations as the other side.

    The bank statement gives one side away free; everything here is the
    determination of the other. Debits and credits are derived from account
    type at this single point, rather than stored, so the two cannot drift.
    """
    coa = get_chart_of_accounts()
    entries: list[JournalEntry] = []

    allocations = repos.allocations.list_for_run(run_id)
    by_transaction: dict[int, list] = {}
    for allocation in allocations:
        by_transaction.setdefault(allocation.bank_transaction_id, []).append(allocation)

    for txn_id, rows in by_transaction.items():
        txn = repos.transactions.get(txn_id)
        if txn is None:
            continue

        entry = JournalEntry(
            entry_id=f"JNL-{run_id}-{txn_id}",
            date=txn.txn_date.isoformat(),
            narration=txn.raw_description,
        )

        document = (
            repos.documents.get(rows[0].matched_document_id)
            if rows[0].matched_document_id else None
        )
        reference = document.doc_number if document else txn.bank_reference

        for allocation in rows:
            if allocation.account_id is None:
                # Unresolved lines are excluded rather than dumped into
                # suspense: an entry that cannot be explained should not be
                # posted at all.
                continue

            account = coa.get(allocation.account_id)
            gross = allocation.amount

            # A matched tax invoice is authoritative over our own arithmetic.
            if document is not None and document.tax_amount is not None and len(rows) == 1:
                tax = document.tax_amount
                net = gross - tax
            else:
                net, tax = split_tax(gross, allocation.tax_code)

            if txn.is_inflow:
                # Money in: credit the account that earned it.
                entry.lines.append(JournalLine(
                    date=entry.date, account_code=account.code, account_name=account.name,
                    description=txn.raw_description, debit=Decimal("0.00"), credit=net,
                    tax_code=allocation.tax_code, reference=reference,
                ))
                if tax:
                    entry.lines.append(JournalLine(
                        date=entry.date, account_code="820", account_name="GST Payable",
                        description="GST on sale", debit=Decimal("0.00"), credit=tax,
                        tax_code=allocation.tax_code, reference=reference,
                    ))
            else:
                entry.lines.append(JournalLine(
                    date=entry.date, account_code=account.code, account_name=account.name,
                    description=txn.raw_description, debit=net, credit=Decimal("0.00"),
                    tax_code=allocation.tax_code, reference=reference,
                ))
                if tax:
                    entry.lines.append(JournalLine(
                        date=entry.date, account_code="820", account_name="GST Payable",
                        description="Input tax", debit=tax, credit=Decimal("0.00"),
                        tax_code=allocation.tax_code, reference=reference,
                    ))

        if not entry.lines:
            continue

        # The bank side, which the statement already told us.
        total = sum((line.debit + line.credit for line in entry.lines), Decimal("0"))
        bank = coa.get(BANK_ACCOUNT_CODE)
        entry.lines.append(JournalLine(
            date=entry.date,
            account_code=BANK_ACCOUNT_CODE,
            account_name=bank.name if bank else "Bank",
            description=txn.raw_description,
            # Money in increases the bank asset, which is a debit in the
            # client's books - the mirror of how the bank prints it.
            debit=total if txn.is_inflow else Decimal("0.00"),
            credit=Decimal("0.00") if txn.is_inflow else total,
            tax_code=None,
            reference=reference,
        ))

        entries.append(entry)

    return entries


def journal_csv(entries: list[JournalEntry]) -> str:
    buffer = StringIO()
    writer = csv.writer(buffer)
    writer.writerow([
        "Entry", "Date", "Account Code", "Account Name", "Description",
        "Debit", "Credit", "Tax Code", "Reference",
    ])
    for entry in entries:
        for line in entry.lines:
            writer.writerow([
                entry.entry_id, line.date, line.account_code, line.account_name,
                line.description,
                f"{line.debit:.2f}" if line.debit else "",
                f"{line.credit:.2f}" if line.credit else "",
                line.tax_code or "", line.reference or "",
            ])
    return buffer.getvalue()


def review_rows(repos: Repositories, run_id: int) -> list[dict]:
    """One row per allocation, in the order an accountant would work through."""
    coa = get_chart_of_accounts()
    rows = []

    for allocation in repos.allocations.list_for_run(run_id):
        txn = repos.transactions.get(allocation.bank_transaction_id)
        account = coa.get(allocation.account_id)
        document = (
            repos.documents.get(allocation.matched_document_id)
            if allocation.matched_document_id else None
        )
        rows.append({
            "date": txn.txn_date.isoformat(),
            "description": txn.raw_description,
            "reference": txn.bank_reference or "",
            "money_in": f"{txn.money_in:.2f}" if txn.money_in else "",
            "money_out": f"{txn.money_out:.2f}" if txn.money_out else "",
            "account_code": allocation.account_id or "",
            "account_name": account.name if account else "UNRESOLVED",
            "tax_code": allocation.tax_code or "",
            "status": str(allocation.status),
            "confidence": f"{allocation.confidence:.2f}" if allocation.confidence is not None else "",
            "decided_by": str(allocation.decision_method),
            "supporting_document": document.doc_number if document else "",
            "document_summary": (document.summary or "") if document else "",
            "reasoning": (allocation.reasoning or "").replace("\n", " ")[:400],
            "question": allocation.question or "",
        })

    # Anything needing attention first, then largest amounts, because that is
    # the order in which an accountant's time is best spent.
    priority = {
        AllocationStatus.CLIENT_QUERY: 0,
        AllocationStatus.NEEDS_REVIEW: 1,
        AllocationStatus.APPROVED: 2,
        AllocationStatus.AUTO_POSTED: 3,
    }
    rows.sort(key=lambda r: (
        priority.get(AllocationStatus(r["status"]), 9),
        -float(r["money_out"] or r["money_in"] or 0),
    ))
    return rows


def review_csv(repos: Repositories, run_id: int) -> str:
    rows = review_rows(repos, run_id)
    if not rows:
        return ""
    buffer = StringIO()
    writer = csv.DictWriter(buffer, fieldnames=list(rows[0]))
    writer.writeheader()
    writer.writerows(rows)
    return buffer.getvalue()


def review_xlsx(repos: Repositories, run_id: int, path: Path) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    rows = review_rows(repos, run_id)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Review"

    if not rows:
        workbook.save(path)
        return path

    headers = list(rows[0])
    sheet.append([h.replace("_", " ").title() for h in headers])
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", start_color="DDDDDD")

    shading = {
        "CLIENT_QUERY": "FFE0E0",
        "NEEDS_REVIEW": "FFF6DD",
        "AUTO_POSTED": "E8F5E9",
        "APPROVED": "E8F5E9",
    }
    for row in rows:
        sheet.append([row[h] for h in headers])
        colour = shading.get(row["status"])
        if colour:
            for cell in sheet[sheet.max_row]:
                cell.fill = PatternFill("solid", start_color=colour)

    widths = {
        "description": 42, "reasoning": 70, "question": 60,
        "account_name": 26, "document_summary": 36,
    }
    for index, header in enumerate(headers, start=1):
        sheet.column_dimensions[sheet.cell(1, index).column_letter].width = widths.get(header, 14)
    sheet.freeze_panes = "A2"
    for cell in sheet["N"]:
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    workbook.save(path)
    return path


def client_queries(repos: Repositories, run_id: int) -> list[dict]:
    """The questions to send the client, ready to paste into an email."""
    queries = []
    for allocation in repos.allocations.list_for_run(run_id, AllocationStatus.CLIENT_QUERY):
        txn = repos.transactions.get(allocation.bank_transaction_id)
        queries.append({
            "allocation_id": allocation.id,
            "date": txn.txn_date.isoformat(),
            "description": txn.raw_description,
            "amount": f"{txn.amount:.2f}",
            "direction": "received" if txn.is_inflow else "paid",
            "question": allocation.question or (
                f"We could not determine what the {txn.amount} "
                f"{'received on' if txn.is_inflow else 'paid on'} "
                f"{txn.txn_date:%d %b %Y} relates to. Could you let us know?"
            ),
        })
    return queries


def run_summary(repos: Repositories, run_id: int) -> dict:
    """Headline numbers for a run, including how much needed a person."""
    run = repos.runs.get(run_id)
    statuses = repos.allocations.status_counts(run_id)
    methods = repos.allocations.decision_method_counts(run_id)
    total = sum(statuses.values()) or 1

    entries = build_journal(repos, run_id)
    unbalanced = [e.entry_id for e in entries if not e.balances]

    return {
        "run_id": run_id,
        "client_id": run.client_id,
        "status": str(run.status),
        "transactions": total,
        "by_status": statuses,
        "by_decision_method": methods,
        "auto_post_rate": round(statuses.get("AUTO_POSTED", 0) / total, 3),
        "needs_attention": statuses.get("NEEDS_REVIEW", 0) + statuses.get("CLIENT_QUERY", 0),
        "resolved_without_model": methods.get(str(DecisionMethod.RULE), 0),
        "llm_calls": run.llm_calls,
        "input_tokens": run.input_tokens,
        "output_tokens": run.output_tokens,
        "journal_entries": len(entries),
        "unbalanced_entries": unbalanced,
    }


def export_all(repos: Repositories, run_id: int, out_dir: Path) -> dict[str, Path]:
    out_dir.mkdir(parents=True, exist_ok=True)
    written: dict[str, Path] = {}

    xlsx = out_dir / f"run-{run_id}-review.xlsx"
    review_xlsx(repos, run_id, xlsx)
    written["review_xlsx"] = xlsx

    csv_path = out_dir / f"run-{run_id}-review.csv"
    csv_path.write_text(review_csv(repos, run_id), encoding="utf-8")
    written["review_csv"] = csv_path

    journal_path = out_dir / f"run-{run_id}-journal.csv"
    journal_path.write_text(journal_csv(build_journal(repos, run_id)), encoding="utf-8")
    written["journal_csv"] = journal_path

    queries_path = out_dir / f"run-{run_id}-client-queries.json"
    queries_path.write_text(
        json.dumps(client_queries(repos, run_id), indent=2), encoding="utf-8"
    )
    written["client_queries"] = queries_path

    summary_path = out_dir / f"run-{run_id}-summary.json"
    summary_path.write_text(
        json.dumps(run_summary(repos, run_id), indent=2, default=str), encoding="utf-8"
    )
    written["summary"] = summary_path

    return written
